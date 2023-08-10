import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, ttk
from tkinter import *
import warnings
import threading
import logging
import os
from pathlib import Path
warnings.filterwarnings('ignore')


# class to generate reports
class ReportGenerator:
    def __init__(self, test_file, _sbgd_file, gui):
        self.gui = gui
        self.test_file = test_file
        self._sbgd_file = _sbgd_file
        desktop_path = str(Path.home() / "Desktop")  # getting the user's desktop path
        self.output_folder = os.path.join(desktop_path, "Output")
        os.makedirs(self.output_folder, exist_ok=True)  # creating output folder on desktop
        self.final_output_SGBD = os.path.join(self.output_folder, "SGBD_New.xlsx")  # creates a file SGBD_New.xlsx on your desktop
        self.test_status_new = os.path.join(self.output_folder, "Test_New.xlsx")  # takes path from desktop

    # function for generating SGBD_New.xlsx
    def generate_SGBD(self):
        # self.isGenerating = True
        print('in sgbd')
        input_dtcs = pd.read_excel(self.test_file, sheet_name='DTCs', header=1)  # reads TestStatus.xlsx DTCs sheet
        input_jobs = pd.read_excel(self.test_file, sheet_name='JOBs', header=1)  # reads TestStatus.xlsx JOBs sheet

        output_dtcs = pd.read_excel(self._sbgd_file, sheet_name='DTCs')  # reads SGBD.xlsx DTCs sheet
        output_jobs = pd.read_excel(self._sbgd_file, sheet_name='Jobs')  # reads SGBD.xlsx Jobs sheet
        # output_uwb = pd.read_excel(self._sbgd_file, sheet_name='UWB-ENV')
        columns_from_dtcs = ['Impl.-Status', 'Test-Status Lieferant', 'Test-Status Sublieferant']  # list of columns to copy from DTCs sheet
        columns_from_jobs = ['Impl.-Status', 'Test-Status', 'Implementiert bis']  # list of columns to copy from JOBS sheet

        total = len(columns_from_dtcs) + len(columns_from_jobs)  # calculating total number of columns to compare in both sheets
        work = 0  # declaring a variable work
        for i in columns_from_dtcs:  # iterating through every column in columns_from_dtcs DTCs sheet
            output_dtcs[i] = input_dtcs[i]  # copies the column i from input_dtcs to output_dtcs ith location
            work += 1  # incrementing work by 1
            self.gui.update_progress_bar((100 / total) * work)  # updating the progress bar
            logging.info('Progress: %.2f%%', (100 / total) * work)  # logging the Progress %
        print('exiting dtcs')

        for i in columns_from_jobs:  # iterating through every column in columns_from_dtcs DTCs sheet
            output_jobs[i] = input_jobs[i]    # copies the column i from input_dtcs to output_dtcs ith location
            work += 1   # incrementing work by 1
            self.gui.update_progress_bar((100 / total) * work)  # updating the progress bar
            logging.info('Progress: %.2f%%', (100 / total) * work)  # logging the Progress %
        print('exiting jobs')
        with pd.ExcelWriter(self.final_output_SGBD, engine='xlsxwriter') as writer:  # opens the final_output_SGBD file
            # print('Writing....')
            output_dtcs.to_excel(writer, sheet_name='DTCs', index=False, startrow=0)  # writes the modified values to output_dtcs in DTCs sheet from zeroth row
            output_jobs.to_excel(writer, sheet_name='Jobs', index=False, startrow=0)  # writes the modified values to output_jobs in Jobs sheet from zeroth row
            # output_uwb.to_excel(writer, sheet_name='UWB-ENV', index=False,startrow=0)
        print('Written Successfully!')
        logging.info('Progress: 100%')

    # function for generating TestStatus_New.xlsx
    def new_test_status(self):
        logging.info('Working on DTC Sheet')
        test_file_dtc = pd.read_excel(self.test_file, sheet_name='DTCs', header=1)  # reads DTCs sheet from test_file.xlsx
        new_sgbd_dtcs = pd.read_excel(self.final_output_SGBD, sheet_name='DTCs')  # reads the DTCs sheet from SGBD_New.xlsx
        to_compare_dtc = ['Impl.-Status', 'Test-Status Lieferant', 'Test-Status Sublieferant']  # list of columns to compare

        modified_rows_dtc = (test_file_dtc[to_compare_dtc] != new_sgbd_dtcs[to_compare_dtc])  # creates a DataFrame.checks for modified values in TestStatus.xlsx and SGBD_New.xlsx and creates a data frame of boolean values.

        wb = load_workbook(self.test_file)  # loads the Test_file.xlsx workbook
        ws = wb['DTCs']  # selects the DTCs sheets in the loaded workbook

        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # print('in test')
        for idx, row in modified_rows_dtc.iterrows():  # iterating through each row in modified_rows DataFrame. idx is the current row and row is a pandas series
            for col_name in to_compare_dtc:  # iterating through each column in to_compare_dtc
                if row[col_name]:  # this checks if the value in the current row column is True or not
                    col_index = test_file_dtc.columns.get_loc(col_name) + 1  # getting the index of the col_name. Adding 1 as column is starting from 2
                    entry = ws.cell(row=idx + 3, column=col_index)  # accessing the cell in DTC sheet. +3 because offset is 3
                    entry.value = new_sgbd_dtcs.at[idx, col_name]
                    entry.fill = fill  # applying highlight
        wb.save(self.test_status_new)  # saving the modified workbook
        logging.info('DTC Sheet modified successfully')

        # print('----------------------------------------------')
        logging.info('Working on JOBs Sheet')
        test_file_jobs = pd.read_excel(self.test_status_new, sheet_name='JOBs', header=1)  # reads JOBs sheet from test_file_new.xlsx
        new_sgbd_jobs = pd.read_excel(self.final_output_SGBD, sheet_name='Jobs')  # reads the Jobs sheet from SGBD_New.xlsx
        to_compare_jobs = ['Impl.-Status', 'Test-Status', 'Implementiert bis']  # list of columns to compare

        modified_rows_dtc = (test_file_jobs[to_compare_jobs] != new_sgbd_jobs[to_compare_jobs])  # creates a DataFrame.checks for modified values in TestStatus.xlsx and SGBD_New.xlsx and creates a data frame of boolean values.

        wj = load_workbook(self.test_status_new)    # loads the TestStatus_New.xlsx workbook
        w = wj['JOBs']  # selects the JOBs sheets in the loaded workbook

        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # print('in jobs')

        for idx, row in modified_rows_dtc.iterrows():  # iterating through each row in modified_rows DataFrame. idx is the current row and row is a pandas series
            for col_name in to_compare_jobs:  # iterating through each column in to_compare_dtc
                if row[col_name]:  # this checks if the value in the current row column is True or not
                    col_index = test_file_jobs.columns.get_loc(col_name) + 1  # getting the index of the col_name. Adding 1 as column is starting from 2
                    entry = w.cell(row=idx + 3, column=col_index)  # accessing the cell in DTC sheet. +3 because offset is 3
                    entry.value = new_sgbd_jobs.at[idx, col_name]
                    entry.fill = fill  # applying highlight

        wj.save(self.test_status_new)  # saving the modified workbook
        logging.info('JOBs Sheet modified successfully')
        logging.info('Both Sheets modified successfully')

    # def new_test_jobs(self):
    #     test_file_jobs = pd.read_excel(self.test_file, sheet_name='JOBs', header=1)
    #     new_sgbd_jobs = pd.read_excel(self.final_output_SGBD, sheet_name='Jobs')
    #     to_compare_jobs = ['Impl.-Status', 'Test-Status', 'Implementiert bis']
    #
    #     modified_rows_dtc = (test_file_jobs[to_compare_jobs] != new_sgbd_jobs[to_compare_jobs])
    #
    #     wb = load_workbook(self.test_file)
    #     ws = wb['JOBs']
    #
    #     fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    #     print('in jobs')
    #     # Apply highlighting to modified entries in Test.xlsx DTCs sheet
    #     for idx, row in modified_rows_dtc.iterrows():
    #         for col_name in to_compare_jobs:
    #             if row[col_name]:
    #                 col_index = test_file_jobs.columns.get_loc(col_name) + 1
    #                 entry = ws.cell(row=idx + 3, column=col_index)
    #                 entry.fill = fill
    #
    #     wb.save(self.test_status_new)


# class of GUI
class GUIFrame:
    def __init__(self):
        # print('In GUI')
        # test_status = r'C:\Users\jazibe\PycharmProjects\Excel\Inputs\Teststatus.xlsx'
        # sbgd_file = r'C:\Users\jazibe\PycharmProjects\Excel\Inputs\SGBD.xlsx'
        self.a_instance = None
        self.window = tk.Tk()
        self.window.title("TestStatusReport")
        self.window.geometry('650x200+100+100')  # window view size
        self.window.resizable(width=False, height=False)  # cannot be increased or decreased by the user
        self.path_holder_sgbd = StringVar()
        self.path_holder_test = StringVar()
        self.create_zip_button = None
        self.test_status_button = None
        self.exit_button = None
        self.sgbd_button = None
        self.execute_button = None
        self.progress_bar = None
        self.checkbox_sgbd = None
        self.checkbox_test = None
        self.text_teststatus = None
        self.sgbd_text = None
        self.checkSGBD = IntVar()
        self.checkTestStatus = IntVar()
        self.progress_var = DoubleVar()
        self.log_message = None
        self.create_widgets()

    # function to create view of GUI
    def create_widgets(self):
        # image
        label1 = tk.Label(self.window, text='Image', bg='#ffb6c1', fg='#fff')
        label1.place(x=0, y=0, width=80, height=120)

        # sgbd_new button
        self.create_zip_button = tk.Button(self.window, text='CREATE ZIP', bg='#000', fg='#fff', activebackground='#000', activeforeground='#fff')
        self.create_zip_button.place(x=0, y=120, width=80, height=30)

        # checkboxes
        self.checkbox_sgbd = tk.Checkbutton(self.window, text='Prepare SGBD Report', onvalue=1, offvalue=0, variable=self.checkSGBD)
        self.checkbox_sgbd.place(x=80, y=2, height=20)

        self.checkbox_test = tk.Checkbutton(self.window, text='Prepare TestStatus Report', onvalue=1, offvalue=0, variable=self.checkTestStatus)
        self.checkbox_test.place(x=80, y=30, height=15)

        # test_status button
        self.test_status_button = tk.Button(self.window, text='Load Test Status', bg='#FF34B3', activebackground='#FF34B3', fg='#fff', command=self.open_test_status)
        self.test_status_button.place(x=80, y=51, width=181, height=35)

        # sgbd button
        self.sgbd_button = tk.Button(self.window, text='Load SGBD', bg='#FF34B3', activebackground='#FF34B3', fg='#fff', command=self.open_SGBD)
        self.sgbd_button.place(x=80, y=86, width=181, height=35)

        # exit button
        self.exit_button = tk.Button(self.window, text='Exit', bg='#808080', fg='#fff', activebackground='#808080', command=self.close)
        self.exit_button.place(x=80, y=120, width=181, height=29)

        self.text_teststatus = Text(self.window)
        self.text_teststatus.place(x=261, y=51, width=479, height=35)

        # sgbd holder
        self.sgbd_text = Text(self.window)
        self.sgbd_text.place(x=261, y=86, width=479, height=35)

        # label info
        info_label = tk.Label(text='SELECT THE EITHER SGBD OR TESTSTATUS OPTION TO PREPARE REPORT', bg='green', fg='#fff')
        info_label.place(x=261, y=1)

        # sgbd button holder
        self.execute_button = tk.Button(self.window, text='EXECUTE', bg='#008000', fg='#fff', activebackground='#008000', command=self.execute)
        self.execute_button.place(x=261, y=120, width=389, height=29)

        # progress bar
        self.progress_bar = ttk.Progressbar(self.window, variable=self.progress_var, maximum=100, mode='determinate')
        self.progress_bar.place(x=0, y=149, width=650, height=24)

        self.log_message = tk.Text(self.window, bg='green', fg='#000')
        self.log_message.place(x=0, y=173, width=650, height=27, anchor='nw')

        self.log_message = tk.Text(self.window)
        self.log_message.place(x=0, y=173, width=650, height=27, anchor='nw')

    # function to open file dialog for SGBD files
    def open_SGBD(self):
        file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])  # taking the path of opened file in file_path variable
        if file_path:   # if file_path then only continue
            if not self.path_holder_sgbd.get():  # if path_holder_sgbd is empty then writing to it
                self.path_holder_sgbd.set(file_path)  # writing here
            else:
                self.path_holder_sgbd.set(file_path)  # if hold_file_path has a data then over-riding it
            self.sgbd_text.delete(1.0, 'end')  # clearing the sgbd_text entry
            self.sgbd_text.insert('end', file_path)  # appending sgbd_text here

    # function to open file dialog for TestStatus files
    def open_test_status(self):
        file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])  # taking the path of opened file in file_path variable
        if file_path:  # if file_path then only continue
            if not self.path_holder_test.get():  # if path_holder_test is empty then writing to it
                self.path_holder_test.set(file_path)  # writing here
            else:
                self.path_holder_test.set(file_path)  # if path_holder_test has a data then over-riding it
            self.text_teststatus.delete(1.0, 'end')  # clearing the text_teststatus entry
            self.text_teststatus.insert('end', file_path)  # appending sgbd_text here

    # function to call generate_SGBD() method
    def generate_report_SBGD(self):
        self.a_instance.generate_SGBD()
        self.window.update_idletasks()

    # function for updating progress bar
    def update_progress_bar(self, progress):
        self.progress_var.set(progress)
        self.window.update_idletasks()
        self.execute_button.config(text=f'{progress}%')

    # function to handle execute button
    def execute(self):
        if len(self.path_holder_test.get()) > 0 and len(self.path_holder_sgbd.get()) > 0:
            logging.info('User clicked Execute button')

            if self.a_instance is None:
                self.a_instance = ReportGenerator(self.path_holder_test.get(), self.path_holder_sgbd.get(), self)

            logging.info('Processing...')

            if self.checkSGBD.get() == 1 and self.checkTestStatus.get() == 0:
                logging.info('Generating SGBD Report')
                self.success_log_message_holder('Processing...')
                self.update_progress_bar(70)
                self.update_progress_bar(90)
                sbgd_thread = threading.Thread(target=self.generate_report_SBGD)
                sbgd_thread.start()
                # self.a_instance.generate_SGBD()
                self.window.after(500)
                self.update_progress_bar(100)
                self.success_log_message_holder('SGBD Report Generated')
                logging.info('SGBD Report Generated')

            elif self.checkSGBD.get() == 0 and self.checkTestStatus.get() == 1:
                logging.info('Generating TestStatus Report')
                self.success_log_message_holder('Processing...')
                self.update_progress_bar(70)
                self.update_progress_bar(90)
                self.a_instance.new_test_status()
                self.window.after(500)
                self.update_progress_bar(100)
                self.success_log_message_holder('TestStatus Report Generated')
                self.window.after(500)
                self.success_log_message_holder('TestStatus_New.xlsx saved on your desktop output folder.')
                logging.info('TestStatus Report Generated')

            elif self.checkSGBD.get() == 1 and self.checkTestStatus.get() == 1:
                logging.info('Generating SGBD and TestStatus Reports')
                self.success_log_message_holder('Processing...')
                self.update_progress_bar(70)
                self.a_instance.generate_SGBD()
                self.update_progress_bar(90)
                self.a_instance.new_test_status()
                self.window.after(500)
                self.update_progress_bar(100)
                self.success_log_message_holder('Reports Generated')
                logging.info('SGBD and TestStatus Reports Generated')

            else:
                self.fail_log_message_holder('Please check at least one option')
                logging.warning('No option selected')

        else:
            self.fail_log_message_holder('Please select both SGBD and TestStatus files')
            logging.warning('Files not selected')

    # function for displaying error message
    def fail_log_message_holder(self, msg):
        self.log_message = tk.Text(self.window, bg='red', fg='#000')
        self.log_message.place(x=0, y=173, width=650, height=27, anchor='nw')
        self.log_message.delete(1.0, 'end')
        self.log_message.insert('end', msg + '\n')
        self.window.after(1000)

    # function for displaying success message
    def success_log_message_holder(self, msg):
        self.log_message = tk.Text(self.window, bg='green', fg='#000')
        self.log_message.place(x=0, y=173, width=650, height=27, anchor='nw')
        self.log_message.delete(1.0, 'end')
        self.log_message.insert('end', msg + '\n')
        self.window.after(1000)

    def getPathForTest(self):
        return self.path_holder_test.get()

    # function for running the app
    def run(self):
        self.window.mainloop()

    # function for closing the app when exit button is clicked
    def close(self):
        self.window.quit()


if __name__ == '__main__':
    # x = ReportGenerator(test_status, sbgd_file)
    # x.generate_SGBD()
    # print('****')
    # x.new_test_status()
    logging.basicConfig(filename=r'C:\Users\jazibe\PycharmProjects\Excel\Log\Log.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    y = GUIFrame()
    y.run()
    logging.info('App Closed')
    print('App Closed')
