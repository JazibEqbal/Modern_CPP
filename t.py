import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class DataProcessor:
    def __init__(self, source_file, destination_file):
        self.source_file = source_file
        self.destination_file = destination_file
        self.new_destination_file = "SBD_new.xlsx"

    def copy_and_paste(self):
        # Load source DTC sheet
        source_dtc = pd.read_excel(self.source_file, sheet_name='DTC', header=1)

        # Create a new destination file and write the modified DTC sheet
        with pd.ExcelWriter(self.new_destination_file) as writer:
            source_dtc.to_excel(writer, sheet_name='D', index=False)

    def compare_and_highlight(self):
        # Load original Test.xlsx sheet DTC
        source_dtc = pd.read_excel(self.source_file, sheet_name='DTC', header=1)

        # Load modified SBD_new.xlsx sheet D
        destination_d = pd.read_excel(self.new_destination_file, sheet_name='D', header=0)

        # Find modified rows
        modified_rows = (source_dtc[['B', 'C', 'E']] != destination_d[['B', 'C', 'E']]).any(axis=1)

        # Load SBD_new.xlsx with openpyxl to apply highlighting
        wb = load_workbook(self.new_destination_file)
        ws = wb['D']

        # Define a fill for highlighting
        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Apply highlighting to modified rows
        for idx, is_modified in enumerate(modified_rows, start=2):  # Assuming rows start from 2
            if is_modified:
                for col in range(2, 5):  # Columns B, C, E
                    ws.cell(row=idx, column=col).fill = fill

        # Save the workbook
        wb.save(self.new_destination_file)

        # Save the modified SBD_new.xlsx as Test_new.xlsx
        destination_d.to_excel("Test_new.xlsx", sheet_name='D', index=False)

if __name__ == "__main__":
    source_filename = "Test.xlsx"
    new_destination_filename = "SBD_new.xlsx"

    processor = DataProcessor(source_filename, new_destination_filename)

    # First, copy and paste columns B, C, E
    processor.copy_and_paste()

    # Then, compare and highlight modified rows, and save Test_new.xlsx
    processor.compare_and_highlight()

    print("Columns copied, modified entries highlighted, and saved in Test_new.xlsx!")
